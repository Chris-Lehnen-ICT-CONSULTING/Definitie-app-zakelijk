"""Import v9 ontologisch model vanuit Excel naar ontologie-tabellen (DEF-305).

Leest Begrippen_Compleet_Model_v9.xlsx en importeert:
- 112 begrippen → ontology_terms
- 125 taxonomische (is-a) relaties → ontology_relationships
- 58 niet-taxonomische relaties → ontology_relationships
- 1 ontological_models record als container

Usage:
    .venv/bin/python scripts/import_v9_model.py [--dry-run] [--db-path data/definities.db]
"""

from __future__ import annotations

import argparse
import json
import logging
import sqlite3
import sys
from pathlib import Path

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

XLSX_PATH = Path("_archief/huidige-versie-oud/Begrippen_Compleet_Model_v9.xlsx")
DEFAULT_DB = Path("data/definities.db")

# Categorie-mapping uit sheet "Categorieën uitleg"
CATEGORIE_TO_UFO = {
    "Soort": "Kind",
    "Rol": "Role",
    "Toestand": "Phase",
    "Eigenschap": "Quality",
    "Gebeurtenis": "Event",
    "Verbinder": "Relator",
}

# Golden set: 25 begrippen geselecteerd op diversiteit
# Mix van categorieën, veel relaties, wettelijke grondslagen
GOLDEN_SET = [
    "Entiteit",
    "Persoon",
    "Natuurlijk Persoon",
    "Identiteit",
    "NP-identiteit",
    "Identiteitskenmerk",
    "Biometrisch kenmerk",
    "Biografisch kenmerk",
    "Identiteitsdocument",
    "Identiteitsbewijs",
    "DigiD",
    "Paspoort",
    "Registratie",
    "Strafrechtketendatabank",
    "Strafrechtelijke identiteit",
    "Identificeren",
    "Verifieren",
    "Identiteitsvaststelling",
    "Identiteitsbehandeling",
    "Ketenpartner",
    "Matching Autoriteit",
    "Wachtwoord",
    "PIN",
    "Authenticatiemiddel",
    "Authenticeren",
]


# ── Excel Parsing ───────────────────────────────────────


def parse_begrippen(wb: openpyxl.Workbook) -> list[dict]:
    """Parse sheet 'Begrippen' → lijst van term dicts."""
    ws = wb["Begrippen"]
    begrippen = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        begrippen.append(
            {
                "term_text": str(row[0]).strip(),
                "categorie_6": str(row[1]).strip() if row[1] else None,
                "ufo_categorie": (
                    CATEGORIE_TO_UFO.get(str(row[1]).strip()) if row[1] else None
                ),
                "definitie": str(row[2]).strip() if row[2] else None,
                "wettelijke_basis": (
                    str(row[3]).strip()
                    if row[3] and str(row[3]).strip() != "-"
                    else None
                ),
                "voorbeelden": [
                    str(row[i]).strip()
                    for i in range(4, 7)
                    if row[i] and str(row[i]).strip()
                ],
                "tegenvoorbeelden": [
                    str(row[i]).strip()
                    for i in range(7, 10)
                    if row[i] and str(row[i]).strip()
                ],
                "bron_url": str(row[10]).strip() if len(row) > 10 and row[10] else None,
            }
        )
    return begrippen


def parse_taxonomie(wb: openpyxl.Workbook) -> list[dict]:
    """Parse sheet 'Taxonomie met Verificatie' → lijst van is-a relaties."""
    ws = wb["Taxonomie met Verificatie"]
    relaties = []
    for row in ws.iter_rows(min_row=3, values_only=True):  # skip leeswijzer + header
        if not row[0] or not row[1]:
            continue
        begrip = str(row[0]).strip()
        supertype = str(row[1]).strip()
        if begrip.lower() == "begrip":  # skip header row
            continue
        relaties.append(
            {
                "source": begrip,
                "target": supertype,
                "type": "is_a",
                "verificatie": str(row[3]).strip() if row[3] else None,
                "juridische_bron": str(row[4]).strip() if row[4] else None,
            }
        )
    return relaties


def parse_relaties(wb: openpyxl.Workbook) -> list[dict]:
    """Parse sheet 'Relaties' → lijst van niet-taxonomische relaties."""
    ws = wb["Relaties"]
    relaties = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1] or not row[2]:
            continue
        relaties.append(
            {
                "source": str(row[0]).strip(),
                "type": str(row[1]).strip(),
                "target": str(row[2]).strip(),
                "toelichting": str(row[3]).strip() if row[3] else None,
            }
        )
    return relaties


def parse_wettelijke_grondslagen(wb: openpyxl.Workbook) -> list[dict]:
    """Parse sheet 'Wettelijke grondslagen'."""
    ws = wb["Wettelijke grondslagen"]
    grondslagen = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        grondslagen.append(
            {
                "wet_artikel": str(row[0]).strip(),
                "volledige_naam": str(row[1]).strip() if row[1] else None,
                "onderwerp": str(row[2]).strip() if row[2] else None,
                "relevante_begrippen": str(row[3]).strip() if row[3] else None,
                "bron_url": str(row[4]).strip() if row[4] else None,
            }
        )
    return grondslagen


# ── Database Import ─────────────────────────────────────


def import_model(
    db_path: Path,
    begrippen: list[dict],
    taxonomie: list[dict],
    relaties: list[dict],
    grondslagen: list[dict],
    dry_run: bool = False,
) -> dict:
    """Importeer het v9 model in de ontologie-tabellen.

    Returns:
        Stats dict met counts.
    """
    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA foreign_keys=ON")

    try:
        # 0. Idempotency check (review fix: voorkom duplicaten bij dubbel draaien)
        existing = conn.execute(
            "SELECT id FROM ontological_models "
            "WHERE model_name = ? AND version_number = ?",
            ("Identiteitsbehandeling v9", 9),
        ).fetchone()
        if existing:
            raise RuntimeError(
                f"Model 'Identiteitsbehandeling v9' (versie 9) bestaat al "
                f"(id={existing[0]}). Verwijder het eerst of gebruik een ander versienummer."
            )

        # 1. Maak ontological_models record
        cursor = conn.execute(
            "INSERT INTO ontological_models (model_name, version_number, snapshot_json) "
            "VALUES (?, ?, ?)",
            (
                "Identiteitsbehandeling v9",
                9,
                json.dumps(
                    {
                        "source": str(XLSX_PATH),
                        "begrippen_count": len(begrippen),
                        "taxonomie_count": len(taxonomie),
                        "relaties_count": len(relaties),
                        "grondslagen_count": len(grondslagen),
                        "golden_set": GOLDEN_SET,
                    },
                    ensure_ascii=False,
                ),
            ),
        )
        model_id = cursor.lastrowid
        logger.info("Model aangemaakt: id=%d", model_id)

        # 2. Insert begrippen → ontology_terms
        term_id_map: dict[str, int] = {}
        for b in begrippen:
            cursor = conn.execute(
                "INSERT INTO ontology_terms "
                "(model_id, term_text, categorie_6, ufo_categorie, "
                "classification_confidence, wettelijke_basis) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (
                    model_id,
                    b["term_text"],
                    b["categorie_6"],
                    b["ufo_categorie"],
                    1.0,  # handmatig model = 100% confidence
                    b["wettelijke_basis"],
                ),
            )
            term_id_map[b["term_text"]] = cursor.lastrowid

        logger.info("Begrippen geïmporteerd: %d", len(begrippen))

        # 3. Insert taxonomische relaties (is-a)
        tax_imported = 0
        tax_skipped = 0
        for rel in taxonomie:
            source_id = term_id_map.get(rel["source"])
            target_id = term_id_map.get(rel["target"])
            if source_id is None or target_id is None:
                logger.warning(
                    "Taxonomie overgeslagen: '%s' is-a '%s' (term niet gevonden)",
                    rel["source"],
                    rel["target"],
                )
                tax_skipped += 1
                continue
            conn.execute(
                "INSERT INTO ontology_relationships "
                "(model_id, source_term_id, target_term_id, relationship_type, "
                "confidence_score, inferred_by) "
                "VALUES (?, ?, ?, 'is_a', 1.0, 'import_v9')",
                (model_id, source_id, target_id),
            )
            tax_imported += 1

        logger.info(
            "Taxonomie geïmporteerd: %d (overgeslagen: %d)", tax_imported, tax_skipped
        )

        # 4. Insert niet-taxonomische relaties
        rel_imported = 0
        rel_skipped = 0
        for rel in relaties:
            source_id = term_id_map.get(rel["source"])
            target_id = term_id_map.get(rel["target"])
            if source_id is None or target_id is None:
                logger.warning(
                    "Relatie overgeslagen: '%s' %s '%s' (term niet gevonden)",
                    rel["source"],
                    rel["type"],
                    rel["target"],
                )
                rel_skipped += 1
                continue
            conn.execute(
                "INSERT INTO ontology_relationships "
                "(model_id, source_term_id, target_term_id, relationship_type, "
                "confidence_score, inferred_by) "
                "VALUES (?, ?, ?, ?, 1.0, 'import_v9')",
                (model_id, source_id, target_id, rel["type"]),
            )
            rel_imported += 1

        logger.info(
            "Relaties geïmporteerd: %d (overgeslagen: %d)", rel_imported, rel_skipped
        )

        # 5. Voorbeelden/tegenvoorbeelden bewaren in snapshot
        # (review fix: ontology_terms schema heeft geen kolommen hiervoor,
        # dus bewaren we ze als mapping in de model snapshot)
        voorbeelden_map = {
            b["term_text"]: {
                "voorbeelden": b.get("voorbeelden", []),
                "tegenvoorbeelden": b.get("tegenvoorbeelden", []),
            }
            for b in begrippen
            if b.get("voorbeelden") or b.get("tegenvoorbeelden")
        }

        # 6. Golden set markeren in snapshot
        golden_ids = {
            term: term_id_map[term] for term in GOLDEN_SET if term in term_id_map
        }
        conn.execute(
            "UPDATE ontological_models SET snapshot_json = ? WHERE id = ?",
            (
                json.dumps(
                    {
                        "source": str(XLSX_PATH),
                        "begrippen_count": len(begrippen),
                        "taxonomie_count": tax_imported,
                        "relaties_count": rel_imported,
                        "grondslagen_count": len(grondslagen),
                        "golden_set": GOLDEN_SET,
                        "golden_set_ids": golden_ids,
                        "wettelijke_grondslagen": grondslagen,
                        "voorbeelden": voorbeelden_map,
                    },
                    ensure_ascii=False,
                ),
                model_id,
            ),
        )

        logger.info(
            "Voorbeelden bewaard: %d termen met voorbeelden/tegenvoorbeelden",
            len(voorbeelden_map),
        )

        stats = {
            "model_id": model_id,
            "begrippen": len(begrippen),
            "taxonomie_imported": tax_imported,
            "taxonomie_skipped": tax_skipped,
            "relaties_imported": rel_imported,
            "relaties_skipped": rel_skipped,
            "grondslagen": len(grondslagen),
            "golden_set": len(golden_ids),
        }

        if dry_run:
            conn.rollback()
            logger.info("DRY RUN — geen data opgeslagen")
        else:
            conn.commit()
            logger.info("Import succesvol gecommit")

        return stats

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# ── Main ────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(description="Import v9 ontologisch model")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse en valideer zonder op te slaan",
    )
    parser.add_argument(
        "--db-path",
        type=Path,
        default=DEFAULT_DB,
        help=f"Database pad (default: {DEFAULT_DB})",
    )
    parser.add_argument(
        "--xlsx-path",
        type=Path,
        default=XLSX_PATH,
        help=f"Excel bestand (default: {XLSX_PATH})",
    )
    args = parser.parse_args()

    if not args.xlsx_path.exists():
        logger.error("Excel bestand niet gevonden: %s", args.xlsx_path)
        sys.exit(1)

    if not args.db_path.exists():
        logger.error("Database niet gevonden: %s", args.db_path)
        sys.exit(1)

    logger.info("=== Import v9 Ontologisch Model (DEF-305) ===")
    logger.info("Bron: %s", args.xlsx_path)
    logger.info("Doel: %s", args.db_path)

    wb = openpyxl.load_workbook(str(args.xlsx_path), read_only=True)

    begrippen = parse_begrippen(wb)
    taxonomie = parse_taxonomie(wb)
    relaties = parse_relaties(wb)
    grondslagen = parse_wettelijke_grondslagen(wb)
    wb.close()

    logger.info("")
    logger.info("Geparsed:")
    logger.info("  Begrippen:     %d", len(begrippen))
    logger.info("  Taxonomie:     %d is-a relaties", len(taxonomie))
    logger.info("  Relaties:      %d niet-taxonomisch", len(relaties))
    logger.info("  Grondslagen:   %d", len(grondslagen))

    # Categorieën verdeling
    cats = {}
    for b in begrippen:
        cat = b["categorie_6"] or "Onbekend"
        cats[cat] = cats.get(cat, 0) + 1
    logger.info("  Categorieën:   %s", cats)

    logger.info("")

    stats = import_model(
        db_path=args.db_path,
        begrippen=begrippen,
        taxonomie=taxonomie,
        relaties=relaties,
        grondslagen=grondslagen,
        dry_run=args.dry_run,
    )

    logger.info("")
    logger.info("=== Resultaat ===")
    for k, v in stats.items():
        logger.info("  %-25s %s", k, v)


if __name__ == "__main__":
    main()
